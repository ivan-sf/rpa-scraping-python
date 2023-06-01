import logging
from tkinter import Tk, Label, Button, filedialog, Entry
from tkinter.ttk import Progressbar
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import schedule
import time
from datetime import datetime
import tkinter.messagebox as messagebox
import matplotlib.pyplot as plt


class ProcesoWebScrapingRPA:
    def __init__(self):
        self.archivo_xlsx = None
        self.columna_producto = "A"
        self.columna_valor_xpath = "B"
        self.columna_estado = "C"
        self.driver = None
        self.wb = None

    def cargar_archivo_xlsx(self):
        try:
            self.wb = load_workbook(self.archivo_xlsx)
            self.hoja = self.wb.active
        except Exception as e:
            logging.error("Error al cargar el archivo XLSX: %s", str(e))
            raise

    def obtener_total_productos(self):
        return self.hoja.max_row - 1

    def iniciar_proceso(self):
        try:
            self.cargar_archivo_xlsx()
            total_productos = self.obtener_total_productos()

            # Inicializar el webdriver
            self.driver = webdriver.Chrome('chromedriver.exe')

            lista_precios = []  # Lista para almacenar los precios

            for i in range(2, self.hoja.max_row + 1):
                producto = self.hoja[self.columna_producto + str(i)].value

                self.actualizar_estado_bot(i, "Procesando")
                self.actualizar_progreso(total_productos, i - 1)

                url = f"https://listado.mercadolibre.com.co/{producto}"
                valor_span = self.obtener_valor_span(url)
                self.actualizar_valor_xpath(i, valor_span)

                self.actualizar_estado_bot(i, "Completado")

                # Mostrar el producto y el precio en la pantalla
                self.mostrar_producto_precio(producto, valor_span)

                lista_precios.append(float(valor_span.replace('.', '').replace(',', '.')))

            self.guardar_archivo_xlsx()
            self.cerrar_webdriver()

            crear_grafica_precios(lista_precios)  # Crear la gráfica de barras con los precios

        except Exception as e:
            logging.error("Error en el proceso de web scraping: %s", str(e))

    def actualizar_estado_bot(self, fila, estado):
        self.hoja[self.columna_estado + str(fila)].value = estado

    def actualizar_progreso(self, total_productos, progreso_actual):
        porcentaje = int((progreso_actual / total_productos) * 100)
        barra_progreso["value"] = porcentaje
        ventana.update()

    def obtener_valor_span(self, url):
        try:
            self.driver.get(url)

            # Buscar el valor del producto en el span con la clase "price-tag-fraction"
            valor_span = None
            elemento_span = self.driver.find_element(By.CSS_SELECTOR, "span.price-tag-fraction")
            if elemento_span:
                valor_span = elemento_span.text

            return valor_span

        except Exception as e:
            logging.error("Error en la navegación web: %s", str(e))
            return None

    def actualizar_valor_xpath(self, fila, valor_span):
        self.hoja[self.columna_valor_xpath + str(fila)].value = valor_span

    def guardar_archivo_xlsx(self):
        try:
            self.wb.save(self.archivo_xlsx)
        except Exception as e:
            logging.error("Error al guardar el archivo XLSX: %s", str(e))

    def cerrar_webdriver(self):
        self.driver.quit()

    def mostrar_producto_precio(self, producto, precio):
        etiqueta_producto.config(text=f"Producto: {producto}")
        etiqueta_precio.config(text=f"Precio: {precio}")


def cargar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("Archivo Excel", "*.xlsx")])
    if archivo:
        etiqueta_archivo.config(text="Archivo seleccionado: " + archivo)
        proceso_scraping_rpa.archivo_xlsx = archivo


def iniciar_proceso_scraping():
    proceso_scraping_rpa.iniciar_proceso()
    entry_hora.config(state="normal")


def programar_proceso_scraping():
    hora_programada = entry_hora.get()
    try:
        datetime.strptime(hora_programada, "%H:%M")
        schedule.clear()
        schedule.every().day.at(hora_programada).do(iniciar_proceso_scraping)
        etiqueta_estado_programacion.config(text=f"Programado a las {hora_programada}")
        messagebox.showinfo("Proceso programado", f"El proceso se ha programado a las {hora_programada}.")
        entry_hora.config(state="disabled")
        while True:
            schedule.run_pending()
            time.sleep(1)
    except ValueError:
        etiqueta_estado_programacion.config(text="Hora no válida (formato HH:MM)")


def crear_grafica_precios(precios):
    plt.figure(figsize=(8, 6))
    productos = range(1, len(precios) + 1)
    plt.bar(productos, precios)
    plt.xlabel('Producto')
    plt.ylabel('Precio')
    plt.title('Precios de los productos')
    plt.xticks(productos)
    plt.show()


# Configurar el registro en el archivo de log
logging.basicConfig(filename='log.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Crear la ventana principal
ventana = Tk()
ventana.title("Ejercicio de Web Scraping y RPA")
ventana.geometry("400x400")

# Crear etiquetas y botones
etiqueta_archivo = Label(ventana, text="Archivo seleccionado: ")
etiqueta_archivo.pack()

boton_cargar = Button(ventana, text="Cargar archivo", command=cargar_archivo)
boton_cargar.pack()

etiqueta_progreso = Label(ventana, text="Progreso:")
etiqueta_progreso.pack()

barra_progreso = Progressbar(ventana, length=200, mode="determinate")
barra_progreso.pack()

etiqueta_producto = Label(ventana, text="")
etiqueta_producto.pack()

etiqueta_precio = Label(ventana, text="")
etiqueta_precio.pack()

proceso_scraping_rpa = ProcesoWebScrapingRPA()

boton_iniciar = Button(ventana, text="Iniciar", command=iniciar_proceso_scraping)
boton_iniciar.pack()

etiqueta_programacion = Label(ventana, text="Programar proceso:")
etiqueta_programacion.pack()

entry_hora = Entry(ventana)
entry_hora.pack()

boton_programar = Button(ventana, text="Programar", command=programar_proceso_scraping)
boton_programar.pack()

etiqueta_estado_programacion = Label(ventana, text="")
etiqueta_estado_programacion.pack()

# Iniciar el bucle principal de la interfaz
ventana.mainloop()
